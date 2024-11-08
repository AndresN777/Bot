import sys
import os
# Añadir la carpeta superior a sys.path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from datetime import datetime
import os
import pandas as pd
from openpyxl import load_workbook

def formatear_fecha_inicial_citisalud(): #Devuelve la fecha del primer día del mes
    t = datetime.now()
    if t.month < 10:
        return f"01/0{t.month}/{t.year}"
    else:
        return f"01/{t.month}/{t.year}"

def formatear_link_pdf(text, lab:str): #Formatea link a partir de un texto
    if lab == "citisalud":
        prefijo = 'https://lis.citisalud.com.co:8087'
        return f"{prefijo}{text[6:]}"
    elif lab == "adb":
        prefijo = 'http://www.resultadoslab.com.co/resultadoslab/'
        return f"{prefijo}{text[3:]}"

def obtener_nombre_carpeta_principal(): #Obtiene el nombre de la carpeta en la que está guardado el código
    ruta_actual = os.path.dirname(os.path.join(os.path.dirname(__file__))) # Obtener ruta del directorio
    nombre_carpeta = os.path.basename(ruta_actual) # Extraer el nombre de la carpeta
    return nombre_carpeta

def formatear_fecha_tamara(fecha):
    text = f"{fecha}"
    return [text, [int(text[8:10]), int(text[5:7]), int(text[:4])]]

def verificar_pagina_tamara(fecha):
    """if not f"{fecha}"[:4].isdigit():
        return 0"""
    año = int(f"{fecha}"[:4])
    año_actual = int(f"{datetime.now().year}")
    if año < año_actual:
        return False
    else:
        return True

def verificar_fecha_tamara(fecha):
    if not f"{fecha}"[:4].isdigit():
        return False
    else:
        aux = formatear_fecha_tamara(fecha)
        if int(datetime.now().year) > aux[1][2]:
            return True
        
        elif int(datetime.now().year) == aux[1][2]:

            if int(datetime.now().month) > aux[1][1]:
                return True
            
            elif int(datetime.now().month) == aux[1][1]:

                if int(datetime.now().day) > aux[1][0]:
                    return True
                else:
                    return False
            else:
                return False
        else:
            return False

def contar_filas_excel(archivo_excel, nombre_hoja=None):
    libro = load_workbook(archivo_excel, data_only=True)
    
    hoja = libro[nombre_hoja] if nombre_hoja else libro.active
    
    num_filas = hoja.max_row
    return num_filas
